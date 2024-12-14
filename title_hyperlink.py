import openpyxl
import subprocess
input_filename = r"hyperlink_test\\sample.xlsx"
output_filename = r"hyperlink_test\\sample_output.xlsx"
# 新しいワークブックを作成
wb = openpyxl.Workbook()
ws = wb.active
ws2 = wb.create_sheet("Sheet2")
# 元のExcelファイルを読み込む
source_wb = openpyxl.load_workbook(input_filename)
source_ws = source_wb["Sheet1"]
# 行を1から末尾まで処理
for row in source_ws.iter_rows(min_row=1, max_row=source_ws.max_row, min_col=1, max_col=2):
    cell_value_A = row[0].value  # A列の値
    cell_value_B = row[1].value  # B列の値
    if cell_value_A and cell_value_B:  # 値がある場合に処理
        dest_row = row[0].row + 1  # 書き込み先の行番号
        ws[f"A{dest_row}"].value = cell_value_A
        ws[f"A{dest_row}"].hyperlink = cell_value_B
# 保存
wb.save(output_filename)
# 出力ファイルを開く
subprocess.Popen(["start", "", output_filename], shell=True)